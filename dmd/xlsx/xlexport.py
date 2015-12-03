#!/usr/bin/env python
# -*- coding: utf-8 -*-
# vim: ai ts=4 sts=4 et sw=4 nu

from __future__ import (unicode_literals, absolute_import,
                        division, print_function)
import logging
import copy
import StringIO

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.dimensions import RowDimension, ColumnDimension
from openpyxl.styles import (PatternFill, Border, Side,
                             Alignment, Protection, Font)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.fills import FILL_SOLID

from dmd.xlsx import column_to_letter
from dmd.models.Entities import Entity
from dmd.models.Indicators import Indicator

logger = logging.getLogger(__name__)


def xl_col_width(cm):
    """ xlwt width for a given width in centimeters """
    return 4.725 * cm


def xl_set_col_width(sheet, column, cm):
    """ change column width """
    letter = column_to_letter(column)
    if column not in sheet.column_dimensions.keys():
        sheet.column_dimensions[letter] = \
            ColumnDimension(worksheet=sheet)
    sheet.column_dimensions[letter].width = xl_col_width(cm)


def xl_row_height(cm):
    """ xlwt height for a given height in centimeters """
    return 28.35 * cm


def xl_set_row_height(sheet, row, cm):
    """ change row height """
    if row not in sheet.row_dimensions.keys():
        sheet.row_dimensions[row] = RowDimension(worksheet=sheet)
    sheet.row_dimensions[row].height = xl_row_height(cm)


dataentry_fname_for = lambda dps: "saisie-PNLP-{}.xlsx".format(
    dps.std_name if dps != Entity.get_root() else "DPS")


def generate_dataentry_for(dps, save_to=None):

    is_all_dps = dps == Entity.get_root()

    # colors
    black = 'FF000000'
    dark_gray = 'FFA6A6A6'
    light_gray = 'FFDEDEDE'
    yellow = 'F9FF00'

    # styles
    header_font = Font(
        name='Calibri',
        size=12,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color=black)

    std_font = Font(
        name='Calibri',
        size=12,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color=black)

    header_fill = PatternFill(fill_type=FILL_SOLID, start_color=dark_gray)
    yellow_fill = PatternFill(fill_type=FILL_SOLID, start_color=yellow)
    black_fill = PatternFill(fill_type=FILL_SOLID, start_color=black)
    odd_fill = PatternFill(fill_type=FILL_SOLID, start_color=light_gray)

    thin_black_side = Side(style='thin', color='FF000000')
    thick_black_side = Side(style='thick', color='FF000000')

    std_border = Border(
        left=thin_black_side,
        right=thin_black_side,
        top=thin_black_side,
        bottom=thin_black_side,
    )

    thick_left_border = Border(
        left=thick_black_side,
        right=thin_black_side,
        top=thin_black_side,
        bottom=thin_black_side,)
    thick_right_border = Border(
        right=thick_black_side,
        left=thin_black_side,
        top=thin_black_side,
        bottom=thin_black_side,)

    centered_alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0)

    left_alignment = Alignment(
        horizontal='left',
        vertical='center')

    vertical_alignment = Alignment(
        horizontal='left',
        vertical='bottom',
        text_rotation=90,
        wrap_text=True,
        shrink_to_fit=False,
        indent=0)

    number_format = '# ### ### ##0'

    protected = Protection(locked=True, hidden=False)
    unprotected = Protection(locked=False, hidden=False)

    header_style = {
        'font': header_font,
        'fill': header_fill,
        'border': std_border,
        'alignment': centered_alignment,
        'protection': protected
    }

    vheader_style = {
        'font': std_font,
        'alignment': vertical_alignment,
        'protection': protected
    }

    vheader_left_style = copy.copy(vheader_style)
    vheader_left_style.update({'border': thick_left_border})
    vheader_right_style = copy.copy(vheader_style)
    vheader_right_style.update({'border': thick_right_border})

    std_style = {
        'font': std_font,
        'border': std_border,
        'alignment': centered_alignment,
    }

    names_style = {
        'font': std_font,
        'border': std_border,
        'alignment': left_alignment,
    }

    def apply_style(target, style):
        for key, value in style.items():
            setattr(target, key, value)

    # data validations
    yv = DataValidation(type="list",
                        formula1='"{}"'.format(
                            ",".join([str(y) for y in range(2014, 2025)])),
                        allow_blank=True)
    mv = DataValidation(type="list",
                        formula1='"{}"'.format(
                            ",".join([str(y) for y in range(1, 13)])),
                        allow_blank=True)
    dv = DataValidation(type="whole", operator="greaterThanOrEqual",
                        formula1='0')

    wb = Workbook()
    ws = wb.active
    ws.title = "Données"
    ws.freeze_panes = ws['C5']

    ws.add_data_validation(yv)
    ws.add_data_validation(mv)
    ws.add_data_validation(dv)

    # resize row height for 0, 1
    xl_set_row_height(ws, 1, 2.2)
    xl_set_row_height(ws, 2, 2.2)

    # resize col A, B
    xl_set_col_width(ws, 1, 5.5)
    xl_set_col_width(ws, 2, 4.5)

    # write partial metadata headers
    ws.merge_cells("A3:A4")
    ws.cell("A3").value = "DPS"

    ws.merge_cells("B3:B4")
    ws.cell("B3").value = "ZS"

    ws.cell("C3").value = "ANNÉE"
    ws.cell("D3").value = "MOIS"

    indicator_column = 5
    dps_row = 5
    # zs_row = dps_row + 1

    # header style
    for sr in openpyxl.utils.cells_from_range("A3:D4"):
        for coord in sr:
            apply_style(ws.cell(coord), header_style)
    for coord in ["C4", "D4"]:
        ws.cell(coord).fill = yellow_fill
        ws.cell(coord).protection = unprotected

    # ZS of the selected DPS
    children = [child for child in dps.get_children()
                if child.etype == Entity.ZONE or is_all_dps]

    def std_write(row, column, value, style=std_style):
        cell = ws.cell(row=row, column=column)
        cell.value = value
        apply_style(cell, style)

    # write indicator headers
    column = indicator_column
    for indicator in Indicator.get_all_manual():

        # write top header with indic name
        row = 1
        ws.merge_cells(start_row=row, end_row=row + 1,
                       start_column=column, end_column=column + 1)
        std_write(row, column, indicator.name, vheader_style)

        # write header with indic number
        row = 3
        num_str = "{n} - {t}".format(n=indicator.number,
                                     t=indicator.verbose_collection_type)
        ws.merge_cells(start_row=row, end_row=row,
                       start_column=column, end_column=column + 1)
        std_write(row, column, num_str, header_style)
        apply_style(ws.cell(row=row, column=column + 1), header_style)

        # write sub header with NUM/DENOM
        row = 4
        if indicator.itype == Indicator.NUMBER:
            ws.merge_cells(start_row=row, end_row=row,
                           start_column=column, end_column=column + 1)
            std_write(row, column, "NOMBRE", std_style)

            for r in range(row, row + len(children) + 2):  # DPS + children
                ws.merge_cells(start_row=r, end_row=r,
                               start_column=column, end_column=column + 1)
        else:
            std_write(row, column, "NUMERAT", std_style)
            std_write(row, column + 1, "DÉNOM", std_style)

        row = dps_row + len(children)
        nb_rows = row if is_all_dps else row + 1

        # whether a row displays a ZS or not
        row_is_zs = lambda row: False if is_all_dps else row > dps_row

        # row-specific styles
        for r in range(1, nb_rows):
            left = ws.cell(row=r, column=column)
            right = ws.cell(row=r, column=column + 1)

            # apply default style
            if r >= dps_row:
                apply_style(left, std_style)
                apply_style(right, std_style)
                left.number_format = number_format
                right.number_format = number_format

                # apply even/odd style
                if r % 2 == 0:
                    if column == indicator_column:
                        for c in range(1, indicator_column):
                            ws.cell(row=r, column=c).fill = odd_fill
                    ws.cell(row=r, column=column).fill = odd_fill
                    ws.cell(row=r, column=column + 1).fill = odd_fill

                # disable cell if data not expected at ZS
                if row_is_zs(r) and indicator.collection_level != Entity.ZONE:
                    left.fill = black_fill
                    left.protection = protected
                    right.fill = black_fill
                    right.protection = protected
                elif not row_is_zs(r) \
                        and indicator.collection_type == indicator.ROUTINE:
                    left.fill = black_fill
                    left.protection = protected
                    right.fill = black_fill
                    right.protection = protected
                else:
                    left.protection = unprotected
                    right.protection = unprotected

            # apply thick borders
            left.border = thick_left_border
            right.border = thick_right_border

        # iterate over indicator
        column += 2

    last_row = dps_row + len(children)

    # apply data validation for periods
    yv.ranges.append('C4:C{}'.format(last_row))
    mv.ranges.append('D4:D{}'.format(last_row))

    # apply positive integer validation to all cells
    last_column = indicator_column + len(Indicator.get_all_manual())
    last_letter = column_to_letter(last_column)
    dv.ranges.append('E4:{c}{r}'.format(c=last_letter, r=last_row))

    row = dps_row
    initial_row = [] if is_all_dps else [None]
    # write names & periods
    for child in initial_row + children:
        if is_all_dps:
            dps_name = child.std_name
            zs_name = "-"
        else:
            dps_name = dps.std_name
            zs_name = child.std_name if child else "-"
        std_write(row, 1, dps_name, names_style)
        std_write(row, 2, zs_name, names_style)

        # set default value for period
        year = ws.cell(row=row, column=3)
        year.set_explicit_value(value="=$C$4",
                                data_type=year.TYPE_FORMULA)
        apply_style(year, std_style)
        year.protection = unprotected

        month = ws.cell(row=row, column=4)
        month.set_explicit_value(value="=$D$4",
                                 data_type=month.TYPE_FORMULA)
        apply_style(month, std_style)
        month.protection = unprotected

        row += 1

    ws.protection.set_password("PNLP")
    ws.protection.enable()

    if save_to:
        logger.info("saving to {}".format(save_to))
        wb.save(save_to)
        return

    stream = StringIO.StringIO()
    wb.save(stream)

    return stream


def export_to_spreadsheet(qs, save_to=None):

    # colors
    black = 'FF000000'
    dark_gray = 'FFA6A6A6'

    # styles
    header_font = Font(
        name='Calibri',
        size=12,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color=black)

    std_font = Font(
        name='Calibri',
        size=12,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color=black)

    header_fill = PatternFill(fill_type=FILL_SOLID, start_color=dark_gray)

    thin_black_side = Side(style='thin', color='FF000000')

    std_border = Border(
        left=thin_black_side,
        right=thin_black_side,
        top=thin_black_side,
        bottom=thin_black_side,
    )

    centered_alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0)

    number_format = '# ### ### ##0'

    header_style = {
        'font': header_font,
        'fill': header_fill,
        'border': std_border,
        'alignment': centered_alignment,
    }

    std_style = {
        'font': std_font,
        'border': std_border,
        'alignment': centered_alignment,
        'number_format': number_format,
    }

    empty = ""

    col_year = 1
    col_month = 2
    col_dps = 3
    col_zs = 4
    col_numerator = 5
    col_denominator = 6
    col_value = 7
    col_human = 8

    def apply_style(target, style):
        for key, value in style.items():
            setattr(target, key, value)

    wb = Workbook()
    wb.remove_sheet(wb.active)

    logger.info("exporting {} records".format(qs.count()))

    # one sheet per indicator
    for indicator in Indicator.objects.all():
        ws = wb.create_sheet()
        ws.title = "#{}".format(indicator.number)

        def std_write(row, column, value, style=std_style):
            cell = ws.cell(row=row, column=column)
            cell.value = value
            apply_style(cell, style)

        row = 1

        # write header
        std_write(row, col_year, "Année", header_style)
        std_write(row, col_month, "Mois", header_style)
        std_write(row, col_dps, "DPS", header_style)
        xl_set_col_width(ws, col_dps, 7)
        std_write(row, col_zs, "ZS", header_style)
        xl_set_col_width(ws, col_zs, 7)
        std_write(row, col_numerator, "Numérateur", header_style)
        xl_set_col_width(ws, col_numerator, 3)
        std_write(row, col_denominator, "Dénominateur", header_style)
        xl_set_col_width(ws, col_denominator, 3)
        std_write(row, col_value, "Valeur", header_style)
        std_write(row, col_human, "Valeur (affich.)", header_style)
        xl_set_col_width(ws, col_human, 3)

        row += 1

        for record in qs.filter(indicator=indicator).iterator():
            logger.debug(record)

            std_write(row, col_year, record.period.year, std_style)
            std_write(row, col_month, record.period.month, std_style)
            std_write(row, col_dps,
                      getattr(record.entity.get_dps(), 'short_name', empty),
                      std_style)
            std_write(row, col_zs,
                      getattr(record.entity.get_zs(), 'short_name', empty),
                      std_style)
            std_write(row, col_numerator, record.numerator, std_style)
            std_write(row, col_denominator, record.denominator, std_style)
            std_write(row, col_value, record.value, std_style)
            std_write(row, col_human, record.human(), std_style)

            row += 1

    if save_to:
        logger.info("saving to {}".format(save_to))
        wb.save(save_to)
        return

    stream = StringIO.StringIO()
    wb.save(stream)

    return stream


def indicators_list_to_spreadsheet(qs, save_to=None):

    # colors
    black = 'FF000000'
    dark_gray = 'FFA6A6A6'

    # styles
    header_font = Font(
        name='Calibri',
        size=12,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color=black)

    std_font = Font(
        name='Calibri',
        size=12,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color=black)

    header_fill = PatternFill(fill_type=FILL_SOLID, start_color=dark_gray)

    thin_black_side = Side(style='thin', color='FF000000')

    std_border = Border(
        left=thin_black_side,
        right=thin_black_side,
        top=thin_black_side,
        bottom=thin_black_side,
    )

    centered_alignment = Alignment(
        horizontal='center',
        vertical='center',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0)

    left_alignment = Alignment(
        horizontal='left',
        vertical='center')

    number_format = '# ### ### ##0'

    header_style = {
        'font': header_font,
        'fill': header_fill,
        'border': std_border,
        'alignment': centered_alignment,
    }

    std_style = {
        'font': std_font,
        'border': std_border,
        'alignment': centered_alignment,
        'number_format': number_format,
    }

    name_style = {
        'font': std_font,
        'border': std_border,
        'alignment': left_alignment,
        'number_format': number_format,
    }

    col_number = 1
    col_name = 2
    col_type = 3
    col_origin = 4

    def apply_style(target, style):
        for key, value in style.items():
            setattr(target, key, value)

    wb = Workbook()
    ws = wb.active
    ws.title = "Indicateurs"

    logger.info("exporting {} records".format(qs.count()))

    def std_write(row, column, value, style=std_style):
        cell = ws.cell(row=row, column=column)
        cell.value = value
        apply_style(cell, style)

    # write header
    row = 1
    std_write(row, col_number, "#", header_style)
    xl_set_col_width(ws, col_number, 1.6)
    std_write(row, col_name, "Indicateur", header_style)
    xl_set_col_width(ws, col_name, 37.3)
    std_write(row, col_type, "Type", header_style)
    std_write(row, col_origin, "Origine", header_style)

    # one sheet per indicator
    for indicator in Indicator.objects.all():

        row += 1

        std_write(row, col_number, indicator.number, std_style)
        std_write(row, col_name, indicator.name, name_style)
        std_write(row, col_type,
                  indicator.verbose_collection_type.encode('utf-8'), std_style)
        std_write(row, col_origin,
                  indicator.verbose_origin.encode('utf-8'), std_style)

    if save_to:
        logger.info("saving to {}".format(save_to))
        wb.save(save_to)
        return

    stream = StringIO.StringIO()
    wb.save(stream)

    return stream
