#!/usr/bin/env python
# -*- coding: utf-8 -*-
# vim: ai ts=4 sts=4 et sw=4 nu

from __future__ import (unicode_literals, absolute_import,
                        division, print_function)
import logging

from django.utils import timezone
from django.utils.translation import ugettext as _
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import load_workbook

from dmd.models import MonthPeriod, Entity, Indicator
from dmd.xlsx import letter_to_column

logger = logging.getLogger(__name__)


class ExcelValueMissing(ValueError):
    pass


class ExcelValueError(ValueError):
    pass


class IncorrectExcelFile(ValueError):
    pass


class UploadPermissionDenied(ValueError):
    pass


def read_xls(filepath, partner):

    if not partner.can_upload:
        raise UploadPermissionDenied(_("{user} is not allowed to submit data")
                                     .format(user=partner))

    try:
        wb = load_workbook(filepath, data_only=True)
    except InvalidFileException:
        raise IncorrectExcelFile(_("Not a proper XLSX Template."))

    ws = wb.active

    nb_rows = len(ws.rows)
    cd = lambda row, column: ws.cell(row=row, column=column).value

    # data holder
    data = {
        'errors': [],
    }

    # record now's time to compare with delays
    submited_on = timezone.now()

    def add_error(row, column=None, indicator=None, error=None, text=None):
        data['errors'].append({
            'row': row,
            'column': column,
            'indicator': Indicator.get_or_none(indicator.slug)
            if indicator else None,
            'slug': error,
            'text': text,
        })

    # retrieve and store default year/month
    default_year_addr = "=$C$4"
    default_month_addr = "=$D$4"
    try:
        default_year = int(float(cd(4, 3)))
    except:
        default_year = None

    try:
        default_month = int(float(cd(4, 4)))
    except:
        default_month = None

    for row in range(5, nb_rows + 1):

        row_has_data = sum([1 for x in ws.rows[row - 1][4:] if x.value])
        if not row_has_data:
            continue

        rdc = Entity.get_root()

        dps = Entity.find_by_stdname(cd(row, 1), parent=rdc)
        if dps is None:
            logger.warning("No DPS for row #{}".format(row))
            continue  # no DPS, no data

        zs = Entity.find_by_stdname(cd(row, 2), parent=dps)
        if zs is None:
            if cd(row, 2).lower().strip() != "-":
                logger.warning("No ZS for row #{}".format(row))
                continue  # no ZS, no data
            else:
                entity = dps
        else:
            entity = zs

        # check upload location authorization
        if partner.upload_location not in entity.get_ancestors():
            add_error(row, error='permission_denied',
                      text=_("You can't submit data for "
                             "that location ({entity})")
                      .format(entity=entity))
            continue

        # retrieve period
        year_str = cd(row, 3)
        if year_str == default_year_addr:
            year = default_year
        else:
            try:
                year = int(float(year_str))
            except:
                year = None

        month_str = cd(row, 4)
        if month_str == default_month_addr:
            month = default_month
        else:
            try:
                month = int(float(month_str))
            except:
                month = None

        if year is None or month is None:
            logger.warning("No year or month for row #{}".format(row))
            add_error(row, error='incorrect_period',
                      text=_("Missing year or month"))
            continue

        try:
            period = MonthPeriod.get_or_create(year, month)
        except ValueError as e:
            logger.warning("Unable to retrieve period: {}".format(e))
            add_error(row, error='incorrect_period',
                      text=_("Unable to retrieve period"))
            continue

        for idx, cell in enumerate(ws.rows[2][4:]):
            if idx % 2 != 0:
                continue  # skip empty merged cols

            column = letter_to_column(cell.column)

            try:
                number = cell.value.split('-')[0].strip()
            except:
                # the header doesn't respect the format
                # better fail everything
                raise IncorrectExcelFile(_("Not a proper XLSX Template."))

            num = cd(row, column)
            denom = cd(row, column + 1)

            # skip if missing numerator
            if not num:
                logger.debug("No numerator for indic #{}".format(number))
                continue

            indicator = Indicator.get_by_number(number)
            # not an expected number
            if indicator is None:
                logger.warning("No indicator found at col #{}".format(column))
                add_error(row, column=cell.column,
                          error='incorrect_indicator',
                          text=_("Unable to match an Indicator"))
                continue

            # check submission period for that Indicator
            if not indicator.can_submit_on(on=submited_on, period=period):
                logger.warning("{on} is not a valid submission time "
                               "for {ind} {period}"
                               .format(on=submited_on, ind=indicator,
                                       period=period))
                add_error(row, column=cell.column,
                          indicator=indicator,
                          error='outside_submussion_delay',
                          text=_("{on} is outside submission period "
                                 "for Indicator #{ind} at {period}")
                          .format(on=submited_on.strftime('%d-%m-%Y'),
                                  ind=indicator.number,
                                  period=period))
                continue

            if not indicator.is_number and denom is None:
                logger.warning("No denominator for indic #{}".format(number))
                add_error(row, column=cell.column,
                          error='missing_denominator',
                          text=_("Missing denominator on "
                                 "non-number Indicator"))
                continue
            elif indicator.is_number:
                denom = 1

            try:
                num = float(num)
                denom = float(denom)
            except:
                add_error(row, column=cell.column,
                          error='incorrect_value',
                          indicator=indicator,
                          text=_("Incorrect value for numerator "
                                 "or denominator `{num} / {denom}`")
                                .format(num=num, denom=denom))
                continue

            ident = "{period}_{slug}".format(period=period.strid,
                                             slug=indicator.slug)
            data.update({ident: {
                'slug': indicator.slug,
                'period': period,
                'entity': entity,
                'numerator': num,
                'denominator': denom}})

    return data
